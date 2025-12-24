import openpyxl
path = "/data/uploads/1_20fcdc0640e9bfb656820ad4810e978adbba6c4927c2035e82c4ffd564b1633b.xlsx"
wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
print("sheets:", wb.sheetnames)
ws = wb["ВДЦ"] if "ВДЦ" in wb.sheetnames else wb[wb.sheetnames[0]]
print("using:", ws.title)
hdr = [ws.cell(row=1, column=c).value for c in range(1, 40)]
print("row1:", hdr)
for r in range(1, 10):
    row = [ws.cell(row=r, column=c).value for c in range(1, 40)]
    if any(isinstance(v, str) and "наимен" in v.lower() for v in row) or any(isinstance(v, str) and "идентификатор" in v.lower() for v in row):
        print("header_row_guess:", r, row)
        break
date_like = []
for c in range(1, 80):
    v = ws.cell(row=1, column=c).value
    if hasattr(v, "year"):
        date_like.append((c, v))
print("date_headers_in_row1:", date_like[:5])