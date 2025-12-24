import datetime as dt
import hashlib
import re
from typing import Any, Iterable

from openpyxl.utils.datetime import from_excel

RU_MONTHS = {
    "ЯНВАРЬ": 1,
    "ФЕВРАЛЬ": 2,
    "МАРТ": 3,
    "АПРЕЛЬ": 4,
    "МАЙ": 5,
    "ИЮНЬ": 6,
    "ИЮЛЬ": 7,
    "АВГУСТ": 8,
    "СЕНТЯБРЬ": 9,
    "ОКТЯБРЬ": 10,
    "НОЯБРЬ": 11,
    "ДЕКАБРЬ": 12,
}

def file_sha256(path: str) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()

def norm_str(v: Any) -> str | None:
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        return s if s else None
    return str(v).strip()

def to_date(v: Any) -> dt.date | None:
    if v is None:
        return None
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    if isinstance(v, (int, float)) and v > 30000:  # excel serial
        try:
            d = from_excel(v)
            if isinstance(d, dt.datetime):
                return d.date()
            return d
        except Exception:
            return None
    if isinstance(v, str):
        s=v.strip()
        # try ISO
        for fmt in ("%Y-%m-%d", "%d.%m.%Y"):
            try:
                return dt.datetime.strptime(s, fmt).date()
            except Exception:
                pass
    return None

def is_month_name(v: Any) -> bool:
    if not isinstance(v, str):
        return False
    return v.strip().upper() in RU_MONTHS

def month_start(year: int, month: int) -> dt.date:
    return dt.date(year, month, 1)

def detect_last_used_col(ws, max_rows: int = 30, cap: int = 400) -> int:
    last = 1
    for r in range(1, max_rows+1):
        for c in range(1, cap+1):
            v = ws.cell(row=r, column=c).value
            if v is not None and v != "":
                last = max(last, c)
    return last

def header_to_index(header: list[Any]) -> dict[str, int]:
    m = {}
    for i, v in enumerate(header, start=1):
        if isinstance(v, str):
            m[v.strip()] = i
    return m

def find_row_with_text(ws, text: str, max_rows: int = 50, col: int = 1) -> int | None:
    t = text.lower()
    for r in range(1, max_rows+1):
        v = ws.cell(row=r, column=col).value
        if isinstance(v, str) and t in v.lower():
            return r
    return None

def iter_rows_values(ws, min_row: int, max_row: int, max_col: int):
    for r in ws.iter_rows(min_row=min_row, max_row=max_row, max_col=max_col, values_only=True):
        yield list(r)
