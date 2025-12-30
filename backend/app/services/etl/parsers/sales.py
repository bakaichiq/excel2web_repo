import datetime as dt
import re
from typing import Tuple

import openpyxl
from openpyxl.utils.datetime import from_excel
import pandas as pd

from app.services.etl.utils import RU_MONTHS, detect_last_used_col, is_month_name, month_start
from app.services.etl.validators import ValidationError


MONTH_NAMES = set(RU_MONTHS.keys())
MONTH_ABBR = {
    "янв": 1,
    "фев": 2,
    "мар": 3,
    "апр": 4,
    "май": 5,
    "июн": 6,
    "июл": 7,
    "авг": 8,
    "сен": 9,
    "сент": 9,
    "окт": 10,
    "ноя": 11,
    "дек": 12,
}


def _upper(v):
    return v.strip().upper() if isinstance(v, str) else None


def _parse_month_cell(raw: object) -> dt.date | None:
    if isinstance(raw, dt.datetime):
        return month_start(raw.date().year, raw.date().month)
    if isinstance(raw, dt.date):
        return month_start(raw.year, raw.month)
    if isinstance(raw, (int, float)) and raw > 30000:
        try:
            d = from_excel(raw)
            if isinstance(d, dt.datetime):
                d = d.date()
            if isinstance(d, dt.date):
                return month_start(d.year, d.month)
        except Exception:
            return None
    if not isinstance(raw, str):
        return None
    s = raw.strip().lower().replace(",", ".")
    m = re.search(r"(янв|фев|мар|апр|май|июн|июл|авг|сен|сент|окт|ноя|дек)[а-я]*\.?\s*(\d{2,4})", s)
    if not m:
        return None
    month = MONTH_ABBR.get(m.group(1))
    if not month:
        return None
    y = int(m.group(2))
    if y < 100:
        y = 2000 + y
    return month_start(y, month)


def _find_month_row(ws, max_rows: int = 30, max_cols: int | None = None) -> int | None:
    max_c = max_cols or ws.max_column
    for r in range(1, max_rows + 1):
        for c in range(1, max_c + 1):
            v = ws.cell(r, c).value
            if is_month_name(v) or _parse_month_cell(v):
                return r
    return None


def _find_header_row(ws, keywords: tuple[str, ...], max_rows: int = 30, max_cols: int | None = None) -> int | None:
    max_c = max_cols or ws.max_column
    for r in range(1, max_rows + 1):
        for c in range(1, max_c + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str):
                s = v.lower()
                if any(k in s for k in keywords):
                    return r
    return None


def _collect_year_by_col(ws, start_row: int, end_row: int, max_cols: int) -> dict[int, int]:
    year_by_col: dict[int, int] = {}
    for r in range(start_row, end_row + 1):
        for c in range(1, max_cols + 1):
            v = ws.cell(r, c).value
            if isinstance(v, int):
                year_by_col[c] = v
            elif isinstance(v, float) and v.is_integer():
                year_by_col[c] = int(v)
            elif isinstance(v, str):
                m = re.search(r"(20\d{2})", v)
                if m:
                    year_by_col[c] = int(m.group(1))
    last_year = None
    for c in range(1, max_cols + 1):
        if c in year_by_col:
            last_year = year_by_col[c]
        elif last_year:
            year_by_col[c] = last_year
    return year_by_col


def _collect_scenario_by_col(ws, start_row: int, end_row: int, max_cols: int) -> dict[int, str]:
    scen_by_col: dict[int, str] = {}
    for r in range(start_row, end_row + 1):
        for c in range(1, max_cols + 1):
            v = ws.cell(r, c).value
            if not isinstance(v, str):
                continue
            s = v.strip().lower()
            if "факт" in s or "продано" in s:
                scen_by_col[c] = "fact"
            elif "план" in s:
                scen_by_col[c] = "plan"
            elif "прогноз" in s:
                scen_by_col[c] = "forecast"
    return scen_by_col


def parse_sales(path: str) -> tuple[pd.DataFrame, list[ValidationError]]:
    errors: list[ValidationError] = []
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sheet_name = None
    for name in wb.sheetnames:
        if name.strip().lower() == "план продаж":
            sheet_name = name
            break
    if not sheet_name:
        return pd.DataFrame(), [ValidationError("Не найден лист 'план продаж'", sheet="план продаж")]

    ws = wb[sheet_name]
    max_cols = detect_last_used_col(ws, max_rows=30, cap=400)

    month_row = _find_month_row(ws, max_rows=30, max_cols=max_cols)
    if not month_row:
        return pd.DataFrame(), [ValidationError("Не найдены месячные колонки в листе 'план продаж'", sheet=sheet_name)]

    header_row = _find_header_row(
        ws,
        keywords=("наименование", "название", "позиция", "объект", "продукт", "площад"),
        max_rows=month_row,
        max_cols=max_cols,
    )
    if not header_row:
        header_row = max(1, month_row - 1)

    year_by_col = _collect_year_by_col(ws, header_row, month_row, max_cols)
    scen_by_col = _collect_scenario_by_col(ws, header_row, month_row, max_cols)

    month_cols = []
    for c in range(1, max_cols + 1):
        raw = ws.cell(month_row, c).value
        mv = _upper(raw)
        parsed = _parse_month_cell(raw)
        if mv in MONTH_NAMES:
            y = year_by_col.get(c)
            if not y:
                if isinstance(raw, str):
                    m = re.search(r"(20\d{2})", raw)
                    if m:
                        y = int(m.group(1))
            if not y:
                continue
            month = month_start(y, RU_MONTHS[mv])
        elif parsed:
            month = parsed
        else:
            continue

        scenario = scen_by_col.get(c, "plan")
        if scenario == "forecast":
            continue
        month_cols.append({"col": c, "month": month, "scenario": scenario})

    if not month_cols:
        return pd.DataFrame(), [ValidationError("Не найдены месячные колонки в листе 'план продаж'", sheet=sheet_name)]

    name_col = 1
    for c in range(1, max_cols + 1):
        v = ws.cell(header_row, c).value
        if isinstance(v, str):
            s = v.strip().lower()
            if any(k in s for k in ("наименование", "название", "позиция", "объект", "продукт", "площад")):
                name_col = c
                break

    data = []
    for r in range(month_row + 1, ws.max_row + 1):
        raw_name = ws.cell(r, name_col).value
        if raw_name is None:
            continue
        name = str(raw_name).strip()
        if not name:
            continue
        name_lower = name.lower()
        if name_lower.startswith("итого") or name_lower.startswith("всего") or name_lower.startswith("сумма"):
            continue

        unit_val = ws.cell(r, name_col + 1).value
        unit = str(unit_val).strip().lower() if unit_val is not None else ""
        if unit and "м2" not in unit and "м²" not in unit:
            # берём только строки в м2 для плана продаж
            continue

        scenario_row = "plan"
        if "факт" in name_lower or "реаль" in name_lower:
            scenario_row = "fact"
        elif "план" in name_lower:
            scenario_row = "plan"
        else:
            # если не можем определить сценарий, пропускаем
            continue

        for mc in month_cols:
            v = ws.cell(r, mc["col"]).value
            if v is None:
                continue
            try:
                val = float(v)
            except Exception:
                continue
            if val == 0:
                continue
            data.append(
                {
                    "item_name": name,
                    "month": mc["month"],
                    "scenario": scenario_row,
                    "area_m2": float(val),
                }
            )

    return pd.DataFrame(data), errors
