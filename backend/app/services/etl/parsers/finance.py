import datetime as dt
import re
from typing import Tuple
import pandas as pd
import openpyxl

from app.services.etl.validators import ValidationError
from app.services.etl.utils import RU_MONTHS, month_start

MONTH_NAMES = set(RU_MONTHS.keys())

def _upper(v):
    return v.strip().upper() if isinstance(v,str) else None

def _find_header_row(ws, keyword: str, max_rows: int = 30) -> int | None:
    kw=keyword.lower()
    for r in range(1, max_rows+1):
        for c in range(1, 6):
            v=ws.cell(r,c).value
            if isinstance(v,str) and kw in v.lower():
                return r
    return None

def _parse_month_columns(ws, year_row: int, month_row: int) -> list[dict]:
    """Return list of dicts: {col:int, month:date, scenario:str}"""
    max_col = ws.max_column
    year_by_col={}
    for c in range(1, max_col+1):
        y=ws.cell(year_row, c).value
        if isinstance(y,int):
            year_by_col[c]=y
        elif isinstance(y,float) and y.is_integer():
            year_by_col[c]=int(y)

    # forward-fill year across columns
    last_year = None
    for c in range(1, max_col+1):
        if c in year_by_col:
            last_year = year_by_col[c]
        elif last_year:
            year_by_col[c] = last_year

    # detect forecast marker per year: column where month_row contains 'ПРОГНОЗ'
    forecast_marker={}
    for c in range(1, max_col+1):
        mv=_upper(ws.cell(month_row,c).value)
        if mv and "ПРОГНОЗ" in mv:
            y=year_by_col.get(c)
            if not y:
                m = re.search(r"(20\d{2})", mv)
                if m:
                    y = int(m.group(1))
            if y:
                forecast_marker[y]=c

    month_cols=[]
    for c in range(1, max_col+1):
        mv=_upper(ws.cell(month_row, c).value)
        if mv in MONTH_NAMES:
            y=year_by_col.get(c)
            if not y:
                # try parse year from nearby month-row text
                raw = ws.cell(month_row, c).value
                if isinstance(raw, str):
                    m = re.search(r"(20\d{2})", raw)
                    if m:
                        y = int(m.group(1))
            if not y:
                continue
            scen="plan"
            mkr=forecast_marker.get(y)
            if mkr and c>mkr:
                scen="forecast"
            month_cols.append({"col": c, "month": month_start(y, RU_MONTHS[mv]), "scenario": scen})
    return month_cols


def _find_month_row(ws, start_row: int, max_rows: int = 8) -> int | None:
    for r in range(start_row, start_row + max_rows):
        for c in range(1, ws.max_column + 1):
            mv = _upper(ws.cell(r, c).value)
            if mv in MONTH_NAMES:
                return r
    return None


def _collect_year_by_col(ws, start_row: int, end_row: int) -> dict[int, int]:
    year_by_col: dict[int, int] = {}
    for r in range(start_row, end_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, int):
                year_by_col[c] = v
            elif isinstance(v, float) and v.is_integer():
                year_by_col[c] = int(v)
            elif isinstance(v, str):
                m = re.search(r"(20\d{2})", v)
                if m:
                    year_by_col[c] = int(m.group(1))
    # forward-fill
    last_year = None
    for c in range(1, ws.max_column + 1):
        if c in year_by_col:
            last_year = year_by_col[c]
        elif last_year:
            year_by_col[c] = last_year
    return year_by_col


def _collect_scenario_by_col(ws, start_row: int, end_row: int) -> dict[int, str]:
    scen_by_col: dict[int, str] = {}
    for r in range(start_row, end_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if not isinstance(v, str):
                continue
            s = v.strip().lower()
            if "прогноз" in s:
                scen_by_col[c] = "forecast"
            elif "факт" in s:
                scen_by_col[c] = "fact"
            elif "план" in s:
                scen_by_col[c] = "plan"
    return scen_by_col

def parse_bdr(path: str) -> tuple[pd.DataFrame, list[ValidationError]]:
    errors=[]
    wb=openpyxl.load_workbook(path, data_only=True, read_only=True)
    if "БДР" not in wb.sheetnames:
        return pd.DataFrame(), [ValidationError("Не найден лист 'БДР'", sheet="БДР")]
    ws=wb["БДР"]
    header_row=_find_header_row(ws, "Статья БДР")  # row containing keyword
    if not header_row:
        return pd.DataFrame(), [ValidationError("Не найдена строка заголовка 'Статья БДР'", sheet="БДР")]
    month_row = _find_month_row(ws, header_row, max_rows=6)
    if not month_row:
        return pd.DataFrame(), [ValidationError("Не найдены месячные колонки в БДР", sheet="БДР")]

    year_by_col = _collect_year_by_col(ws, header_row, month_row)
    scen_by_col = _collect_scenario_by_col(ws, header_row, month_row)

    month_cols=[]
    for c in range(1, ws.max_column + 1):
        mv=_upper(ws.cell(month_row, c).value)
        if mv in MONTH_NAMES:
            y=year_by_col.get(c)
            if not y:
                raw = ws.cell(month_row, c).value
                if isinstance(raw, str):
                    m = re.search(r"(20\d{2})", raw)
                    if m:
                        y = int(m.group(1))
            if not y:
                continue
            scenario = scen_by_col.get(c, "plan")
            if scenario == "forecast":
                continue
            month_cols.append({"col": c, "month": month_start(y, RU_MONTHS[mv]), "scenario": scenario})
    if not month_cols:
        return pd.DataFrame(), [ValidationError("Не найдены месячные колонки в БДР", sheet="БДР")]

    # detect name column
    name_col = 1
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if isinstance(v, str) and "статья" in v.lower() and "бдр" in v.lower():
            name_col = c
            break

    data=[]
    start_row=month_row+1
    current_parent: str | None = None
    for r in range(start_row, ws.max_row+1):
        raw = ws.cell(r, name_col).value
        name = raw
        if name is None:
            continue
        if isinstance(name,str) and name.strip()=="":
            continue
        account = str(name).strip()
        indent = 0
        if isinstance(raw, str):
            indent = len(raw) - len(raw.lstrip())
        parent_name = None
        if indent == 0:
            current_parent = account
        else:
            parent_name = current_parent
        for mc in month_cols:
            v=ws.cell(r, mc["col"]).value
            if v is None:
                continue
            try:
                val=float(v)
            except Exception:
                continue
            if val==0:
                continue
            data.append({
                "account_name": account,
                "parent_name": parent_name,
                "month": mc["month"],
                "scenario": mc["scenario"],
                "amount": val,
            })
    return pd.DataFrame(data), errors

def parse_bdds(path: str) -> tuple[pd.DataFrame, list[ValidationError]]:
    errors=[]
    wb=openpyxl.load_workbook(path, data_only=True, read_only=True)
    if "БДДС" not in wb.sheetnames:
        return pd.DataFrame(), [ValidationError("Не найден лист 'БДДС'", sheet="БДДС")]
    ws=wb["БДДС"]
    header_year=_find_header_row(ws, "Статья БДДС") or _find_header_row(ws, "Статья БДДС".lower())
    if not header_year:
        # in this file headers start at row3
        header_year=3
    header_month=header_year+1
    month_cols=_parse_month_columns(ws, header_year, header_month)
    if not month_cols:
        return pd.DataFrame(), [ValidationError("Не найдены месячные колонки в БДДС", sheet="БДДС")]

    data=[]
    start_row=header_month+1
    current_parent=None
    for r in range(start_row, ws.max_row+1):
        name=ws.cell(r,1).value
        if name is None:
            continue
        if isinstance(name,str) and name.strip()=="":
            continue
        account=str(name).strip()

        # Heuristic: treat rows without indentation as parent when they end with 'деятельность' or are category totals
        if isinstance(name,str) and ("деятельност" in name.lower()):
            current_parent=account
        parent_name=current_parent if current_parent!=account else None

        for mc in month_cols:
            v=ws.cell(r, mc["col"]).value
            if v is None:
                continue
            try:
                val=float(v)
            except Exception:
                continue
            if val==0:
                continue
            direction = "in" if val>=0 else "out"
            data.append({
                "account_name": account,
                "parent_name": parent_name,
                "month": mc["month"],
                "scenario": mc["scenario"],
                "direction": direction,
                "amount": val,
            })
    return pd.DataFrame(data), errors
