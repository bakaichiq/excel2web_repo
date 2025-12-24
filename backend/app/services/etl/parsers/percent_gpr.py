import datetime as dt
import pandas as pd
import openpyxl

from app.services.etl.utils import to_date, norm_str

def parse_percent_gpr(xlsx_path: str, sheet_name: str = "%ГПР") -> tuple[pd.DataFrame, pd.DataFrame]:
    """Returns (weights_df, totals_df).

    weights_df columns: ugpr, month(date), weight(0..1)
    totals_df columns: ugpr, plan_total (float|None)
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    if sheet_name not in wb.sheetnames:
        cand = [s for s in wb.sheetnames if "ГПР" in s and "%" in s]
        if not cand:
            return (
                pd.DataFrame(columns=["ugpr","month","weight"]),
                pd.DataFrame(columns=["ugpr","plan_total"]),
            )
        sheet_name = cand[0]
    ws = wb[sheet_name]

    header = [ws.cell(row=2, column=c).value for c in range(1, 300)]
    date_cols = []
    for idx, v in enumerate(header, start=1):
        d = to_date(v)
        if d:
            date_cols.append((idx, dt.date(d.year, d.month, 1)))

    rows=[]
    totals=[]
    for r in range(3, ws.max_row+1):
        ugpr = norm_str(ws.cell(row=r, column=1).value)
        if not ugpr:
            continue
        if ugpr.lower().startswith("итого"):
            break
        plan_total = ws.cell(row=r, column=5).value
        try:
            plan_total_f = float(plan_total) if plan_total not in (None,"") else None
        except Exception:
            plan_total_f = None
        totals.append({"ugpr": ugpr, "plan_total": plan_total_f})
        for c_idx, m in date_cols:
            v = ws.cell(row=r, column=c_idx).value
            if v is None or v == "":
                continue
            try:
                w = float(v)
            except Exception:
                continue
            if w == 0:
                continue
            if w > 1.5:
                w = w/100.0
            rows.append({"ugpr": ugpr, "month": m, "weight": w})
    return pd.DataFrame(rows), pd.DataFrame(totals).drop_duplicates(subset=["ugpr"], keep="last")
