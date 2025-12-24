import datetime as dt
import pandas as pd
import openpyxl

from app.core.config import settings
from app.services.etl.utils import to_date, norm_str

def parse_resources_daily(xlsx_path: str, sheet_name: str = "Люди техника") -> pd.DataFrame:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    if sheet_name not in wb.sheetnames:
        cand=[s for s in wb.sheetnames if "Люди" in s or "техника" in s.lower()]
        if not cand:
            return pd.DataFrame(columns=["resource_name","category","date","scenario","qty","manhours"])
        sheet_name=cand[0]
    ws=wb[sheet_name]
    header=[ws.cell(row=1,column=c).value for c in range(1,400)]
    date_cols=[]
    for idx,v in enumerate(header,start=1):
        d=to_date(v)
        if d:
            date_cols.append((idx,d))
    if not date_cols:
        return pd.DataFrame(columns=["resource_name","category","date","scenario","qty","manhours"])

    rows=[]
    for r in range(2, ws.max_row+1):
        name=norm_str(ws.cell(row=r,column=1).value)
        if not name: 
            continue
        cat=norm_str(ws.cell(row=r,column=2).value) or "люди"
        scenario_raw=norm_str(ws.cell(row=r,column=4).value) or "факт"
        scenario="fact" if "факт" in scenario_raw.lower() else "plan" if "план" in scenario_raw.lower() else "fact"
        for c_idx,d in date_cols:
            v=ws.cell(row=r,column=c_idx).value
            if v is None or v=="":
                continue
            try:
                qty=float(v)
            except Exception:
                continue
            if qty==0:
                continue
            manhours=None
            if cat.lower().startswith("люд") or cat.lower().startswith("персон") or cat.lower() in ("people","human"):
                manhours=qty*float(settings.SHIFT_HOURS)
            rows.append({"resource_name":name,"category":cat,"date":d,"scenario":scenario,"qty":qty,"manhours":manhours})
    return pd.DataFrame(rows)
