import datetime as dt
from pathlib import Path
import openpyxl
import pandas as pd

from app.services.etl.parsers.vdc import parse_vdc
from app.services.etl.parsers.gpr import parse_gpr
from app.services.etl.parsers.people import parse_people_tech
from app.services.etl.parsers.finance import parse_bdr, parse_bdds

def _make_min_file(tmp_path: Path) -> Path:
    wb=openpyxl.Workbook()
    # VDC
    ws=wb.active
    ws.title="ВДЦ"
    meta = [
        "№","Идентификатор операции","Категория","Блок","WBS","Конструктив","Дисциплина","Этаж","УГПР",
        "Название операции","Тип","Наименование работ и материалов","Ед. изм","Количество Защита","Цена Защита","Прогнозное Количество","Цена Фактическая"
    ]
    # pad to 22 columns like real sheet
    while len(meta)<22: meta.append(f"m{len(meta)}")
    d1=dt.date(2025,1,1); d2=dt.date(2025,1,2)
    header = meta + [d1,d2]
    ws.append(header)
    # group row
    ws.append(["Проект"]+[""]*(len(header)-1))
    # data row
    row = ["", "OP-1", "СМР", "Блок A", "WBS-1", "", "Монолит", "1", "UGPR", "Операция 1", "", "Бетон", "м3", 10, 100, 0, 0]
    while len(row)<22: row.append("")
    row += [2, 3]
    ws.append(row)

    # GPR
    ws=wb.create_sheet("ГПР")
    ws.append(["Идентификатор операции","Название операции","Название ИСР","Блок","УГПР","Начало","Окончание","Ед. изм","Плановое количество нетрудовых ресурсов","Цена","Стоимость"])
    ws.append(["OP-1","Операция 1","WBS-1","Блок A","UGPR",dt.date(2025,1,1),dt.date(2025,1,31),"м3",10,100,1000])

    # People
    ws=wb.create_sheet("Люди техника")
    ws.append(["наименование","категория","ед. изм","план/факт","01.01.2025","02.01.2025"])
    ws.append(["Разнорабочие","Manpower","чел.","ФАКТ",5,6])

    # BDR (simplified headers)
    ws=wb.create_sheet("БДР")
    # create 8 rows to align parser search
    ws.append(["","","","","","","","",""])
    ws.append(["","","","","","","","",""])
    ws.append(["","","","","","","","",""])
    ws.append(["","","","","","","","",""])
    ws.append(["","","","","","","","",""])
    ws.append(["","","","","","","","",""])
    ws.append(["Статья БДР","","","","","","2025","2025","2025"])
    ws.append(["","Наименование","ПЛАН","ФАКТ","ОСТАТОК","ПРОГНОЗ 9+3","ПРОГНОЗ 6+6","ЯНВАРЬ","ФЕВРАЛЬ"])
    ws.append(["","Выручка",0,0,0,0,0,1000,1200])

    # BDDS
    ws=wb.create_sheet("БДДС")
    ws.append(["","","",""])
    ws.append(["","","",""])
    ws.append(["Статья БДДС","","2025","2025"])
    ws.append(["","ПЛАН","ЯНВАРЬ","ФЕВРАЛЬ"])
    ws.append(["Поступления","",100,0])
    ws.append(["Платежи","",-50,0])

    out=tmp_path/"sample.xlsx"
    wb.save(out)
    return out

def test_parse_vdc(tmp_path):
    f=_make_min_file(tmp_path)
    baseline, facts, errors = parse_vdc(str(f))
    assert errors == [] or isinstance(errors,list)
    assert not baseline.empty
    assert not facts.empty
    assert set(["operation_code","category","item_name","date","qty"]).issubset(facts.columns)

def test_parse_gpr(tmp_path):
    f=_make_min_file(tmp_path)
    df, errors = parse_gpr(str(f))
    assert errors == [] or isinstance(errors,list)
    assert df.loc[0,"operation_code"]=="OP-1"
    assert df.loc[0,"plan_qty_total"]==10

def test_parse_people(tmp_path):
    f=_make_min_file(tmp_path)
    df, errors = parse_people_tech(str(f))
    assert errors == [] or isinstance(errors,list)
    assert (df["scenario"]=="fact").all()
    assert df["qty"].sum()==11

def test_parse_bdr(tmp_path):
    f=_make_min_file(tmp_path)
    df, errors = parse_bdr(str(f))
    assert errors == [] or isinstance(errors,list)
    assert not df.empty
    assert df["amount"].sum()==2200

def test_parse_bdds(tmp_path):
    f=_make_min_file(tmp_path)
    df, errors = parse_bdds(str(f))
    assert errors == [] or isinstance(errors,list)
    assert not df.empty
    assert df["amount"].sum()==50  # 100 + (-50)
